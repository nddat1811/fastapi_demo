from openpyxl import load_workbook
import os
from app.utils.helper import convert_date
from openpyxl.styles.borders import Border, Side
from io import BytesIO

def edit_excel_file(input_data, buffer):
    current_directory = os.path.dirname(os.path.realpath(__file__))

    # Navigate to the parent directory (one level up from the current directory)
    parent_directory = os.path.dirname(current_directory)

    # Construct the path to the Excel file
    excel_file_path = os.path.join(parent_directory, "resources", "templates", "water_bill_template.xlsx")
    wb = load_workbook(excel_file_path)
    ws = wb['Sheet1']

    data =  convert_date(input_data)
    start_row = 2  # Bắt đầu từ hàng thứ 2 (dòng tiêu đề là hàng thứ 1)
    start_col = 1  # Bắt đầu từ cột A

    thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
    
    # Duyệt qua mỗi đối tượng trong data và ghi vào bảng tính Excel
    for i, item in enumerate(data, start=start_row):
        ws.cell(row=i, column=start_col).value = item["username"]
        ws.cell(row=i, column=start_col + 1).value = item["prev_volume"]
        ws.cell(row=i, column=start_col + 2).value = item["cur_volume"]
        ws.cell(row=i, column=start_col + 3).value = item["total_volume"]
        ws.cell(row=i, column=start_col + 4).value = item["price"]
        ws.cell(row=i, column=start_col + 5).value = item["total_volume_price"]
        ws.cell(row=i, column=start_col + 6).value = item["due_date"]
        ws.cell(row=i, column=start_col + 7).value = item["payment_date"]
        ws.cell(row=i, column=start_col + 8).value = item["created_date"]

        for col_num in range(start_col, start_col + 9):
            ws.cell(row=i, column=col_num).border = thin_border

    for letter in ['F', 'G', 'H', 'I']:
        max_width = 0

        for row_number in range(1, ws.max_row + 1):
            cell_value = ws[f'{letter}{row_number}'].value
            if isinstance(cell_value, str) and len(cell_value) > max_width:
                max_width = len(cell_value)

        ws.column_dimensions[letter].width = max_width + 1
    wb.save(buffer)
    buffer.seek(0)