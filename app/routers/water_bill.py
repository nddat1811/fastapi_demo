from io import BytesIO
import os
from typing import List
from fastapi import APIRouter, Depends, HTTPException, Request, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.db.database import get_db
from app.db import db_water_bill
from app.auth.oauth2 import RoleChecker, get_current_user
from app.models.user import DbUser
from app.schemas.water_bill import NewWaterBillRequest, WaterBillResponse, UpdateWaterBillRequest
import pandas as pd
from fastapi.responses import StreamingResponse, Response
from openpyxl.styles.borders import Border, Side

from app.utils.constants import Role
from app.utils.export_excel import edit_excel_file
from app.utils.helper import convert_date
router = APIRouter(
    prefix='/water-bills',
    tags=['water-bills']
)

@router.post('/new')
async def create_new_water_bills(req: NewWaterBillRequest, 
    db: Session = Depends(get_db), 
    _ : bool = Depends(RoleChecker(allowed_roles=[Role.ADMIN, Role.STAFF])),
    user : DbUser = Depends(get_current_user)
    ):
    water_bill = await db_water_bill.create_water_bill(req, user, db)
    return water_bill


@router.get('/list-bill', response_model = List[WaterBillResponse])
async def list_water_bill(db: Session = Depends(get_db), 
    _ : bool = Depends(RoleChecker(allowed_roles=[Role.ADMIN, Role.STAFF]))
    ):
    bills = await db_water_bill.get_list_water_bill(db)
    return bills


#xem lại router này
@router.get('/user-bills/{user_id}', 
    response_model = List[WaterBillResponse],
    description="Get List Water Bill of user by user id")
async def list_user_water_bill(user_id: int, 
    db: Session = Depends(get_db), 
    _ : bool = Depends(RoleChecker(allowed_roles=[Role.ADMIN, Role.STAFF, Role.USER]))
    ):
    bills = await db_water_bill.get_list_water_bill_by_user_id(user_id, db)
    return bills

#Xem lại router này
@router.get('/detail/{id}', 
    response_model=WaterBillResponse,
    description="Get Water Bill detail by id")
async def get_water_bill_by_id(id: int, 
    db: Session = Depends(get_db), 
    _ : bool = Depends(RoleChecker(allowed_roles=[Role.ADMIN, Role.STAFF, Role.USER]))
    ):
    bill = await db_water_bill.get_water_bill_by_id(id, db)
    return bill

@router.put('/update/{id}', response_model=WaterBillResponse)
async def update_water_bill(req: UpdateWaterBillRequest, 
    id: int, 
    db: Session = Depends(get_db), 
    _ : bool = Depends(RoleChecker(allowed_roles=[Role.ADMIN, Role.STAFF]))
    ):
    bill = await db_water_bill.update_water_bill(req, id, db)
    return bill


@router.get('/export-csv-water-bill-user/{user_id}',
        description="Get water bill of user without template",
        summary="Water bill of user without template"        
        )
async def export_csv(user_id: int, 
    db: Session = Depends(get_db),
    _ : bool = Depends(RoleChecker(allowed_roles=[Role.ADMIN, Role.STAFF, Role.USER]))
    ):
    data: WaterBillResponse = await db_water_bill.get_list_water_bill_by_user_id(user_id, db)
    # Extract relevant data from DbWaterBill objects
    processed_data = []
    for bill in data:
        payment_date = None
        if bill.payment_date is not None:
            payment_date = bill.payment_date.date().isoformat()
        processed_data.append({
            'id': bill.id,
            'username': bill.user.username,
            'creator': bill.creator.username,
            'prev_volume': bill.prev_volume,
            'cur_volume': bill.cur_volume,
            'total_volume': bill.total_volume,
            'price': bill.price,
            'due_date': bill.due_date.date().isoformat(),
            'payment_date': payment_date,
            'total_volume_price': bill.total_volume_price
        })

    # Convert the processed data into a DataFrame
    df = pd.DataFrame(processed_data)
    df.style.set_table_attributes('')

    return StreamingResponse(
        iter([df.to_csv(index=False)]),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename=data.csv"})


@router.get('/export-excel-water-bill-in-month',
        description="Get water bill in input month with template",
        summary="Water bill of 1 month with template"
        )
async def test(m: int=4,
    db: Session = Depends(get_db),
    _ : bool = Depends(RoleChecker(allowed_roles=[Role.ADMIN, Role.STAFF]))
    ):
    bills = await db_water_bill.get_water_bill_in_month(m, db)
    # filename = "resources/templates/water_bill_template.xlsx"
    current_directory = os.path.dirname(os.path.realpath(__file__))

    # Navigate to the parent directory (one level up from the current directory)
    parent_directory = os.path.dirname(current_directory)

    # Construct the path to the Excel file
    excel_file_path = os.path.join(parent_directory, "resources", "templates", "water_bill_template.xlsx")

    # Create buffer to write excel file
    buffer= BytesIO()
    wb = load_workbook(excel_file_path)
    ws = wb['Sheet1']

    data =  convert_date(bills)
    start_row = 2  # Start from row 2 (First row is title)
    start_col = 1  # Start from column A

    thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
    
    # Loop for all items in data and write into Excel
    for i, item in enumerate(data, start=start_row):
        ws.cell(row=i, column=start_col).value = item["username"]
        ws.cell(row=i, column=start_col + 1).value = item["prev_volume"]
        ws.cell(row=i, column=start_col + 2).value = item["cur_volume"]
        ws.cell(row=i, column=start_col + 3).value = item["total_volume"]
        ws.cell(row=i, column=start_col + 4).value = item["price"]
        ws.cell(row=i, column=start_col + 5).value = item["total_volume_price"]
        ws.cell(row=i, column=start_col + 6).value = item["due_date"]
        ws.cell(row=i, column=start_col + 7).value = item["payment_date"]
        ws.cell(row=i, column=start_col + 8).value = item["created_date"]
        ws.cell(row=i, column=start_col + 9).value = item["creator"]

        for col_num in range(start_col, start_col + 10):
            ws.cell(row=i, column=col_num).border = thin_border

    for letter in ['F', 'G', 'H', 'I', 'J']:
        max_width = 0

        for row_number in range(1, ws.max_row + 1):
            cell_value = ws[f'{letter}{row_number}'].value
            if isinstance(cell_value, str) and len(cell_value) > max_width:
                max_width = len(cell_value)

        ws.column_dimensions[letter].width = max_width + 1
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename=water_bill.xlsx"})
